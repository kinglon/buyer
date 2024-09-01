import os.path
import sys
import time
from logutil import LogUtil
from apple import AppleUtil
import requests
import json
import threading
from datetime import datetime
from datamodel import DataModel
import shutil
import openpyxl


# 工作目录
g_work_path = ''

# 购买结果列表
g_buy_results = []

# 当前计划
g_current_plan = {}

# 购买手机型号
g_buy_phone_model = {}

# 同步锁
g_lock = threading.Lock()


class BuyResult:
    def __init__(self):
        # 购买参数，DataModel
        self.buy_param = DataModel()

        # 标志是否成功下单
        self.success = False

        # 订单号
        self.order_number = ''

        # 下单失败原因
        self.fail_reason = ''

        # 开始购买的时间，utc秒数
        self.begin_buy_time = 0

        # 上号的代理地址
        self.proxy_server = ''


class ProxyPool:
    def __init__(self):
        self.proxys = []
        self.next_proxy_index = 0

    def get_next_proxy(self):
        proxy_ip = self.proxys[self.next_proxy_index]['ip']
        proxy_port = self.proxys[self.next_proxy_index]['port']
        proxy_address = "socks5://{}:{}".format(proxy_ip, proxy_port)
        proxy = {"http": proxy_address, "https": proxy_address}
        self.next_proxy_index = (self.next_proxy_index + 1) % len(self.proxys)
        return proxy


# 获取代理IP列表，返回（success, proxy ips, message）
def get_proxy_server_list(proxy_region):
    no_proxy_server = {"http": "", "https": ""}
    error = (False, [], '')
    try:
        url = ("http://api.proxy.ipidea.io/getBalanceProxyIp?big_num=900&return_type=json&lb=1&sb=0&flow=1&regions={}&protocol=socks5"
               .format(proxy_region))
        response = requests.get(url, proxies=no_proxy_server, timeout=10)
        if not response.ok:
            print("获取代理IP列表失败，错误是：{}".format(response))
            return error
        else:
            data = response.content.decode('utf-8')
            data = json.loads(data)
            if data['success']:
                return True, data['data'], ''
            else:
                return True, [], data['msg']
    except Exception as e:
        print("获取代理IP列表失败，错误是：{}".format(e))
        return error


def need_change_proxy(response):
    if response is None:
        return False

    if response.status_code == 403 or response.status_code == 503:
        return True

    return False


def buy(proxys, datamodel):
    try:
        buy_result = BuyResult()
        buy_result.begin_buy_time = datetime.now().timestamp()
        buy_result.buy_param = datamodel

        print('账号({})开始购买'.format(datamodel.account))
        apple_util = AppleUtil()

        proxy_pool = ProxyPool()
        proxy_pool.proxys = proxys
        apple_util.proxies = proxy_pool.get_next_proxy()
        buy_result.proxy_server = apple_util.proxies
        print('使用代理：{}'.format(buy_result.proxy_server))

        # 添加商品
        model = g_buy_phone_model['model']
        code = g_buy_phone_model['code']
        while True:
            print('加入购物包: {}, {}'.format(model, code))
            if apple_util.add_cart(model, code):
                break
            if need_change_proxy(apple_util.last_response):
                apple_util.proxies = proxy_pool.get_next_proxy()
                print('使用代理：{}'.format(apple_util.proxies))

        # 打开购物车
        while True:
            print('打开购物包')
            x_aos_stk = apple_util.open_cart()
            if x_aos_stk:
                break
            if need_change_proxy(apple_util.last_response):
                apple_util.proxies = proxy_pool.get_next_proxy()
                print('使用代理：{}'.format(apple_util.proxies))

        # 进入购物流程
        while True:
            print('进入购物流程')
            ssi = apple_util.checkout_now(x_aos_stk)
            if ssi:
                break
            if need_change_proxy(apple_util.last_response):
                apple_util.proxies = proxy_pool.get_next_proxy()
                print('使用代理：{}'.format(apple_util.proxies))

        # 登录
        while True:
            print('登录账号')
            if apple_util.login(datamodel.account, datamodel.password):
                break
            if need_change_proxy(apple_util.last_response):
                apple_util.proxies = proxy_pool.get_next_proxy()
                print('使用代理：{}'.format(apple_util.proxies))

        # 绑定账号
        while True:
            print('绑定账号')
            pltn = apple_util.bind_account(x_aos_stk, ssi)
            if pltn:
                break
            if need_change_proxy(apple_util.last_response):
                apple_util.proxies = proxy_pool.get_next_proxy()
                print('使用代理：{}'.format(apple_util.proxies))

        # 开始下单
        while True:
            print('开始下单')
            if apple_util.checkout_start(pltn):
                break
            if need_change_proxy(apple_util.last_response):
                apple_util.proxies = proxy_pool.get_next_proxy()
                print('使用代理：{}'.format(apple_util.proxies))

        # 进入下单页面
        while True:
            print('进入下单页面')
            x_aos_stk = apple_util.checkout()
            if x_aos_stk:
                break
            if need_change_proxy(apple_util.last_response):
                apple_util.proxies = proxy_pool.get_next_proxy()
                print('使用代理：{}'.format(apple_util.proxies))

        # 选择自提
        while True:
            print('选择自提')
            if apple_util.fulfillment_retail(x_aos_stk):
                break
            if need_change_proxy(apple_util.last_response):
                apple_util.proxies = proxy_pool.get_next_proxy()
                print('使用代理：{}'.format(apple_util.proxies))

        # 选择店铺
        while True:
            print('选择店铺')
            if apple_util.fulfillment_store(x_aos_stk, datamodel):
                break
            if need_change_proxy(apple_util.last_response):
                apple_util.proxies = proxy_pool.get_next_proxy()
                print('使用代理：{}'.format(apple_util.proxies))

        # 选择联系人
        while True:
            print('选择联系人')
            if apple_util.pickup_contact(x_aos_stk, datamodel):
                break
            if need_change_proxy(apple_util.last_response):
                apple_util.proxies = proxy_pool.get_next_proxy()
                print('使用代理：{}'.format(apple_util.proxies))

        # 输入账单，支付
        while True:
            print('支付')
            if apple_util.billing(x_aos_stk, datamodel):
                break
            if need_change_proxy(apple_util.last_response):
                apple_util.proxies = proxy_pool.get_next_proxy()
                print('使用代理：{}'.format(apple_util.proxies))

        # 确认下单
        while True:
            print('确认订单')
            if apple_util.review(x_aos_stk):
                break
            if need_change_proxy(apple_util.last_response):
                apple_util.proxies = proxy_pool.get_next_proxy()
                print('使用代理：{}'.format(apple_util.proxies))

        # 查询处理结果
        while True:
            time.sleep(2)
            # 处理订单
            print('处理订单')
            finish, sucess = apple_util.query_process_result(x_aos_stk)
            if not finish:
                continue
            break

        buy_result.success = sucess
        if sucess:
            # 查询订单号
            order_no = apple_util.query_order_no()
            if order_no is None:
                print('未查询到订单号')
            else:
                print('订单号是：{}'.format(order_no))
                buy_result.order_number = order_no

        # 保存结果
        g_lock.acquire()
        g_buy_results.append(buy_result)
        save_buy_result_to_json()
        g_lock.release()
    except Exception as e:
        print('遇到问题：{}'.format(e))


# 加载数据，返回(success, buy_time, [datamodel])
def load_data():
    global g_current_plan
    global g_buy_phone_model

    failed_result = (False, 0, [])
    try:
        param_file_path = os.path.join(g_work_path, 'python_args_buy.json')
        with open(param_file_path, 'r', encoding='utf-8') as file:
            json_data = file.read()
            root = json.loads(json_data)
            plan_id = root['plan_id']
            g_buy_phone_model = root["phone"]
            users = root['user']

        current_file_path = os.path.dirname(os.path.abspath(__file__))
        plan_file_path = os.path.join(current_file_path, '..', 'Configs', 'plan.json')
        with open(plan_file_path, 'r', encoding='utf-8') as file:
            json_data = file.read()
            root = json.loads(json_data)
            plans = root['plan']

        g_current_plan = None
        for plan in plans:
            if plan_id == plan['id']:
                g_current_plan = plan
                break
        if g_current_plan is None:
            print('找不到计划')
            return failed_result
        if not g_current_plan['enable_fix_time_buy']:
            print('不是设定时间开始购买')
            return failed_result

        config_file_path = os.path.join(current_file_path, '..', 'Configs', 'configs.json')
        with open(config_file_path, 'r', encoding='utf-8') as file:
            json_data = file.read()
            root = json.loads(json_data)
            shops = root['shop']

        # 获取店铺列表
        current_shops = []
        for plan_shop in g_current_plan['shop']:
            found = False
            for shop in shops:
                if shop['name'] == plan_shop:
                    found = True
                    current_shops.append(shop)
                    break
            if not found:
                print("无法找到店铺")
                return failed_result
        if len(current_shops) == 0:
            print('没有配置店铺')
            return failed_result

        data_models = []
        next_shop = 0
        for user in users:
            data_model = DataModel()
            data_model.account = user['account']
            data_model.state = user['state']
            data_model.password = user['password']
            data_model.city = user['city']
            data_model.email = user['email']
            data_model.first_name = user['first_name']
            data_model.postal_code = user['postal_code']
            data_model.store = current_shops[next_shop]['id']
            data_model.store_name = current_shops[next_shop]['name']
            data_model.store_postal_code = current_shops[next_shop]['postal']
            data_model.street2 = user['street2']
            data_model.street = user['street']
            data_model.telephone = user['telephone']
            data_model.last_name = user['last_name']
            data_model.expired_date = user['expired_date']
            data_model.credit_card_no = user['credit_card_no']
            data_model.cvv = user['cvv']
            data_model.giftcard_no = ''
            data_models.append(data_model)
            next_shop = (next_shop + 1) % len(current_shops)
        return True, g_current_plan['fix_buy_time'], data_models
    except Exception as e:
        print('加载数据遇到问题：{}'.format(e))
        return failed_result


# 保存购买结果到表格文件
def save_buy_result_to_excel():
    # 拷贝模板表格
    current_file_path = os.path.dirname(os.path.abspath(__file__))
    template_file_path = os.path.join(current_file_path, '..', 'Configs', '购买结果.xlsx')
    now = datetime.now()
    output_excel_file_name = now.strftime("购买结果_%Y%m%d_%H%M%S.xlsx")
    output_excel_file = os.path.join(g_work_path, output_excel_file_name)
    if os.path.exists(output_excel_file):
        os.remove(output_excel_file)
        time.sleep(0.3)
    shutil.copy(template_file_path, output_excel_file)

    # 加载表格
    workbook = openpyxl.load_workbook(output_excel_file, data_only=True)
    sheet = workbook.worksheets[0]
    row = 2
    for buy_result in g_buy_results:
        column = 1
        sheet.cell(row=row, column=column, value='')
        column += 1
        sheet.cell(row=row, column=column, value=g_current_plan['name'])
        column += 1
        sheet.cell(row=row, column=column, value=g_current_plan['count'])
        column += 1
        begin_buy_time_string = datetime.fromtimestamp(buy_result.begin_buy_time).strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=row, column=column, value=begin_buy_time_string)
        column += 1
        sheet.cell(row=row, column=column, value=g_buy_phone_model['name'])
        column += 1
        sheet.cell(row=row, column=column, value=str(buy_result.proxy_server))
        column += 1
        sheet.cell(row=row, column=column, value='')
        column += 1
        sheet.cell(row=row, column=column, value=('购物成功' if buy_result.success else '购物失败'))
        column += 1
        sheet.cell(row=row, column=column, value=buy_result.order_number)
        column += 1
        sheet.cell(row=row, column=column, value='')
        column += 1
        order_link = ''
        if len(buy_result.order_number) > 0:
            order_link = ('https://www.apple.com/xc/jp/vieworder/{}/{}'
                          .format(buy_result.order_number, buy_result.buy_param.account))
        sheet.cell(row=row, column=column, value=order_link)
        column += 1
        sheet.cell(row=row, column=column, value='jp')
        column += 1
        buy_param_obj = buy_result.buy_param
        name = buy_param_obj.first_name + buy_param_obj.last_name
        address = buy_param_obj.state + buy_param_obj.city + buy_param_obj.street + buy_param_obj.street2
        order_info = "电话：{}, 名字：{}， 地址：{}".format(buy_param_obj.telephone, name, address)
        sheet.cell(row=row, column=column, value=order_info)
        column += 1
        sheet.cell(row=row, column=column, value=buy_param_obj.account)
        column += 1
        sheet.cell(row=row, column=column, value=buy_param_obj.store_name)
        column += 1
        sheet.cell(row=row, column=column, value='')
        column += 1
        sheet.cell(row=row, column=column, value=buy_param_obj.email)
        column += 1
        sheet.cell(row=row, column=column, value=buy_param_obj.password)
        column += 1
        sheet.cell(row=row, column=column, value=buy_result.fail_reason)
        column += 1

        row += 1

    # 保存修改后的Excel文件
    workbook.save(output_excel_file)


# 保存购买结果到表格文件
def save_buy_result_to_json():
    root = []
    for buy_result in g_buy_results:
        item = {
            'account': buy_result.buy_param.account,
            'success': buy_result.success,
            'order_number': buy_result.order_number
        }
        root.append(item)

    config_file_path = os.path.join(g_work_path, r'buy_result.json')
    with open(config_file_path, "w", encoding='utf-8') as file:
        json.dump(root, file, indent=4)


def main():
    global g_work_path
    g_work_path = sys.argv[1]

    # 启用日志
    LogUtil.set_log_path(g_work_path)
    LogUtil.enable()

    print('加载数据')
    success, buy_time, datamodels = load_data()
    if not success:
        return

    # 获取代理IP列表
    proxy_server_list = []
    message = ''
    for i in range(3):
        print('获取代理IP列表')
        success, proxy_server_list, message = get_proxy_server_list('jp')
        if not success:
            continue
        break

    if len(proxy_server_list) == 0:
        print('获取代理IP列表失败，错误是：{}'.format(message))
        return

    # 倒计时
    last_print_time = 0
    update_proxy = False
    while True:
        time.sleep(0.01)
        now = datetime.now().timestamp()
        if now >= buy_time:
            break
        elapse = int(buy_time - now)
        if now - last_print_time >= 10:  # 每隔10秒打印一次
            print('倒计时：{}秒'.format(elapse))
            last_print_time = now

        # 最后5分钟更新下代理
        if not update_proxy and elapse < 300:
            update_proxy = True
            success, proxy_server_list2, message = get_proxy_server_list('jp')
            if success:
                proxy_server_list = proxy_server_list2
    print('开始购买')

    # 多线程并发购买
    threads = []
    thread_num = len(datamodels)
    proxy_count_per_thread = len(proxy_server_list) // thread_num
    for i in range(thread_num):
        if i < thread_num - 1:
            proxys = proxy_server_list[i * proxy_count_per_thread: (i + 1) * proxy_count_per_thread]
        else:
            proxys = proxy_server_list[i * proxy_count_per_thread:]
        args = (proxys, datamodels[i])
        thread = threading.Thread(target=buy, args=args)
        thread.start()
        threads.append(thread)
    for thread in threads:
        thread.join()

    # 保存购买结果到表格文件
    print('保存购买结果')
    save_buy_result_to_excel()


if __name__ == '__main__':
    main()
    print('购买完成')
    time.sleep(1200)
